#!/usr/bin/env python3
"""
build_report.py — render the consolidated MRP-by-finished-good report.

Reads two JSON files produced by the mrp-by-fg skill during a run:

  --bom  bom_<fg>.json    the exploded, purchased-only component list
  --mrp  mrp_rows_<fg>.json  raw rows kept from the unfiltered planning search
                             "MRP Consolidated Report - Weekly"
                             (only rows whose "Item" is a purchased component)

and writes, into --out-dir:

  MRP_<fg>_<stamp>.xlsx    unstacked grid, one metric per row, weeks across
  MRP_<fg>_<stamp>.html    NetSuite-style stacked dashboard, one page

Pure standard-library + openpyxl. No network, no NetSuite dependency — all
data arrives via the two JSON files, so this file is deterministic and
unit-testable against a captured fixture.

Design notes live in query-notes.md. The saved-search cell format is fixed:
each weekly cell is 7 stacked <div>s in this order —
  [0] Date (or the literal "Overdue" in the Overdue column)
  [1] Forecast
  [2] Demand
  [3] Planned_Demand   (tinted #0F9ED5)
  [4] Supply
  [5] Planned_Supply   (tinted #28a745)
  [6] Ending
A value of "-" means null. Numbers carry thousands separators and padding.
"""

import argparse
import html
import json
import os
import re
import sys

# ---- legend-driven layout --------------------------------------------------
# The planning search stores its per-item row labels in the single-space
# column (" "). Different search variants use different label SETS and ORDER
# (e.g. the 7-field "Ending" layout vs the 9-field layout that adds Supply,
# a separator, and a "System Balance" line). So we read the legend and map
# each cell's stacked values to it BY POSITION — never a hardcoded index.
# Date is always the first row; a blank/separator label (the "-----" divider)
# is skipped.
METRIC_COLOR = {"planned_demand": "0F9ED5", "planned_supply": "28A745"}
# Friendly display names when the normalized key is known; otherwise the
# legend label is title-cased.
_DISPLAY = {
    "forecast": "Forecast", "demand": "Demand", "supply": "Supply",
    "ending": "Ending", "planned_demand": "Planned Demand",
    "planned_supply": "Planned Supply", "system_balance": "System Balance",
}
# Balance-style rows to visually emphasize (the projected on-hand lines).
EMPHASIS = {"ending", "system_balance"}
# Fallback legend if a run has no " " column (classic 7-field layout).
_FALLBACK_KEYS = ["date", "forecast", "demand", "planned_demand", "supply",
                  "planned_supply", "ending"]

RESERVED_KEYS = {"Item", "Description", " ", "Current QTY"}

_DIV_RE = re.compile(r"<div[^>]*>(.*?)</div>", re.S | re.I)
_BR_RE = re.compile(r"<br\s*/?>", re.I)
_TAG_RE = re.compile(r"<[^>]+>")


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------
def _clean(text):
    """Strip tags, unescape entities, collapse whitespace. '' -> None."""
    if text is None:
        return None
    text = _TAG_RE.sub("", text)
    text = html.unescape(text).replace("\xa0", " ").strip()
    return text or None


def _to_number(text):
    """'          12,059.24' -> 12059.24 ; '-' -> None ; '(1,2)' -> -1.2."""
    if text is None:
        return None
    t = text.strip()
    if t in ("", "-", "—") or set(t) == {"-"}:  # '-' or a '-----' separator
        return None
    neg = False
    if t.startswith("(") and t.endswith(")"):
        neg, t = True, t[1:-1]
    t = t.replace(",", "").replace("$", "").strip()
    if t.startswith("-"):
        neg, t = True, t[1:]
    try:
        val = float(t)
    except ValueError:
        return None
    return -val if neg else val


def _norm_key(label):
    """'Planned_Demand' -> 'planned_demand'; 'System Balance' -> 'system_balance'."""
    if not label:
        return None
    k = re.sub(r"[^a-z0-9]+", "_", label.strip().lower()).strip("_")
    return k or None


def _segments(cell_html):
    """Split a cell into its stacked lines regardless of layout: multiple
    <div> children (one per line), OR a single <div> / plain string with <br>
    separators. Returns cleaned strings (entries may be None)."""
    if cell_html is None:
        return []
    divs = _DIV_RE.findall(cell_html)
    if len(divs) > 1:
        raw = divs
    elif len(divs) == 1:
        raw = _BR_RE.split(divs[0])
    else:
        raw = _BR_RE.split(cell_html)
    return [_clean(s) for s in raw]


def legend_keys(mrp_rows):
    """Ordered normalized keys from the row-label legend (the ' ' column).
    keys[i] aligns with segment i of every cell; 'date' marks the date row;
    None marks a blank/separator row (skipped)."""
    legend = None
    for r in mrp_rows:
        if r.get(" "):
            legend = r[" "]
            break
    if not legend:
        return list(_FALLBACK_KEYS)
    return [_norm_key(_clean(s)) for s in _BR_RE.split(legend)]


def metrics_from_keys(keys):
    """[(key, display)] for each non-date, non-blank legend row, in order."""
    out = []
    for k in keys:
        if not k or k == "date":
            continue
        out.append((k, _DISPLAY.get(k, k.replace("_", " ").title())))
    return out


def parse_cell(cell_html, keys):
    """Map a cell's stacked values to the legend keys -> {'date':..., key:num}."""
    vals = _segments(cell_html)
    out = {"date": None}
    for i, key in enumerate(keys):
        v = vals[i] if i < len(vals) else None
        if key == "date":
            out["date"] = None if (v is None or v.lower() == "overdue") else v
        elif key:
            out[key] = _to_number(v)
    return out


def _week_sort_key(col_name):
    n = col_name.strip().lower()
    if n == "overdue":
        return -2
    if n == "this week":
        return -1
    if n == "next week":
        return 0
    m = re.match(r"(\d+)\s+weeks?\s+out", n)
    return int(m.group(1)) if m else 999


def week_columns(mrp_rows):
    """Ordered list of the time-bucket column names present in the data."""
    seen = []
    for row in mrp_rows:
        for k in row.keys():
            if k not in RESERVED_KEYS and k not in seen:
                seen.append(k)
    return sorted(seen, key=_week_sort_key)


def week_dates(mrp_rows, cols, keys):
    """Map column-name -> week-ending date string (from the first row having it)."""
    dates = {}
    for col in cols:
        for row in mrp_rows:
            if col in row:
                d = parse_cell(row[col], keys).get("date")
                if d:
                    dates[col] = d
                    break
        dates.setdefault(col, "Overdue" if col.lower() == "overdue" else "")
    return dates


# ---------------------------------------------------------------------------
# Assemble the model the renderers consume
# ---------------------------------------------------------------------------
def build_model(bom, mrp_rows):
    cols = week_columns(mrp_rows)
    keys = legend_keys(mrp_rows)
    metrics = metrics_from_keys(keys)
    dates = week_dates(mrp_rows, cols, keys)
    # The search keys rows by item INTERNAL ID (its "Item" column), not the
    # displayed itemid. We match on internal_id and display itemid.
    index = {str(r.get("Item")): r for r in mrp_rows}

    # De-dup purchased components, summing qty_per_fg, preserving first order.
    order, comp_map = [], {}
    for c in bom.get("components", []):
        iid = str(c["itemid"])
        if iid not in comp_map:
            comp_map[iid] = {
                "itemid": iid,
                "internal_id": str(c.get("internal_id")) if c.get("internal_id") is not None else None,
                "type": c.get("type"),
                "qty_per_fg": 0.0,
                "level": c.get("level"),
                "parents": [],
            }
            order.append(iid)
        comp_map[iid]["qty_per_fg"] += float(c.get("qty_per_fg") or 0)
        p = c.get("parent")
        if p and p not in comp_map[iid]["parents"]:
            comp_map[iid]["parents"].append(p)

    components, gaps = [], []
    for iid in order:
        cm = comp_map[iid]
        row = index.get(cm["internal_id"]) if cm["internal_id"] else None
        if row is None:
            gaps.append(iid)
            cm["data_gap"] = True
            cm["description"] = None
            cm["current_qty"] = None
            cm["cells"] = {}
        else:
            cm["data_gap"] = False
            desc = (_clean(row.get("Description")) or "").replace("\r\n", " — ").replace("\r", " ").replace("\n", " — ")
            cm["description"] = desc or None
            cq_raw = row.get("Current QTY")
            cm["current_qty"] = _to_number(_clean(str(cq_raw))) if cq_raw not in (None, "") else None
            cm["cells"] = {col: parse_cell(row.get(col), keys) for col in cols}
        components.append(cm)

    return {
        "fg_item": bom.get("fg_item"),
        "fg_description": bom.get("fg_description"),
        "generated_at": bom.get("generated_at"),
        "location": bom.get("location"),
        "search_id": bom.get("search_id", "MRP Consolidated Report"),
        "weeks": cols,
        "week_dates": dates,
        "metrics": metrics,
        "components": components,
        "data_gaps": gaps,
        "wip_assemblies": bom.get("wip_assemblies", []),
        "non_stock_lines": bom.get("non_stock_lines", []),
        "warnings": bom.get("warnings", []),
    }


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def write_xlsx(model, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "MRP by FG"

    INK = "272838"
    MUTED = "64748B"
    GAP_FILL = PatternFill("solid", fgColor="FFE5E9")
    HEAD_FILL = PatternFill("solid", fgColor="272838")
    BAND_FILL = PatternFill("solid", fgColor="F2F2F4")
    thin = Side(style="thin", color="E0E1E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    num_fmt = "#,##0.00;[Red](#,##0.00)"
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    weeks = model["weeks"]
    # Columns: A Component | B Metric | C Current Qty | D Qty/FG | E.. weeks
    fixed = ["Component", "Metric", "Current Qty", "Qty / FG"]
    ncols = len(fixed) + len(weeks)

    # Title block
    ws.cell(1, 1, "Consolidated MRP — Purchased Components").font = Font(bold=True, size=16, color=INK)
    loc = model.get("location") or "All locations"
    ws.cell(2, 1, f'Finished good {model["fg_item"]} — {model.get("fg_description") or ""}').font = Font(size=11, color=INK)
    ws.cell(3, 1, f'Source: NetSuite saved search {model["search_id"]} (Planning Engine Result) · {loc} · generated {model.get("generated_at") or ""}').font = Font(size=9, color=MUTED)
    nwip = len(model.get("wip_assemblies") or [])
    nns = len(model.get("non_stock_lines") or [])
    ns_txt = f' · {nns} non-stock/costing line(s) excluded' if nns else ""
    ws.cell(4, 1, f'Scope: {len(model["components"])} purchased components · full multi-level BOM explosion · {nwip} WIP sub-assembly(ies) traversed (not listed){ns_txt}').font = Font(size=9, color=MUTED)
    if model["data_gaps"]:
        ws.cell(5, 1, f'DATA GAP: {len(model["data_gaps"])} component(s) not found in the planning results — see shaded rows. Values are NOT zero; they are unknown.').font = Font(size=9, bold=True, color="FF0022")

    header_row = 6
    # Header labels
    labels = fixed + [f'{w}\n{model["week_dates"].get(w, "")}' for w in weeks]
    for j, lab in enumerate(labels, start=1):
        c = ws.cell(header_row, j, lab)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = HEAD_FILL
        c.alignment = center
        c.border = border

    metrics = model["metrics"]
    r = header_row + 1
    for comp in model["components"]:
        first = r
        gap = comp["data_gap"]
        for mi, (mkey, mlabel) in enumerate(metrics):
            # Component label only on the first metric row of the block
            comp_lbl = ""
            if mi == 0:
                comp_lbl = comp["itemid"]
                if comp.get("description"):
                    comp_lbl += f'\n{comp["description"]}'
            cc = ws.cell(r, 1, comp_lbl)
            cc.alignment = Alignment(wrap_text=True, vertical="top")
            cc.font = Font(bold=True, color=INK, size=9)

            ws.cell(r, 2, mlabel).font = Font(color=METRIC_COLOR.get(mkey, "1E293B"),
                                              bold=mkey in EMPHASIS, size=9)
            if mi == 0:
                cq = ws.cell(r, 3, comp["current_qty"])
                cq.number_format = num_fmt
                cq.alignment = right
                qf = ws.cell(r, 4, round(comp["qty_per_fg"], 6) if comp["qty_per_fg"] else None)
                qf.number_format = "#,##0.000000"
                qf.alignment = right

            for wj, wcol in enumerate(weeks, start=len(fixed) + 1):
                val = None if gap else comp["cells"].get(wcol, {}).get(mkey)
                cell = ws.cell(r, wj, val)
                cell.number_format = num_fmt
                cell.alignment = right
            # banding + border across the row
            for j in range(1, ncols + 1):
                cell = ws.cell(r, j)
                cell.border = border
                if gap:
                    cell.fill = GAP_FILL
                elif (mi % 2) == 1:
                    cell.fill = BAND_FILL
            r += 1
        # merge the component label vertically across its metric rows
        ws.merge_cells(start_row=first, start_column=1, end_row=r - 1, end_column=1)
        ws.merge_cells(start_row=first, start_column=3, end_row=r - 1, end_column=3)
        ws.merge_cells(start_row=first, start_column=4, end_row=r - 1, end_column=4)

    # widths + freeze
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 12
    for j in range(len(fixed) + 1, ncols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 13
    ws.freeze_panes = ws.cell(header_row + 1, len(fixed) + 1)
    ws.sheet_view.showGridLines = False
    wb.save(path)


# ---------------------------------------------------------------------------
# HTML dashboard
# ---------------------------------------------------------------------------
def _fmt(v):
    if v is None:
        return '<span class="nil">–</span>'
    return f"{v:,.2f}"


def write_html(model, path):
    weeks = model["weeks"]
    wd = model["week_dates"]
    metrics = model["metrics"]

    head_cells = "".join(
        f'<th class="wk"><div class="wk-name">{html.escape(w)}</div>'
        f'<div class="wk-date">{html.escape(wd.get(w, ""))}</div></th>'
        for w in weeks
    )

    body_rows = []
    for comp in model["components"]:
        gap = comp["data_gap"]
        desc = html.escape(comp.get("description") or "") if not gap else ""
        rowspan = len(metrics)
        cls = "gap" if gap else ""
        cq = "" if comp["current_qty"] is None else f"{comp['current_qty']:,.2f}"
        qf = f"{comp['qty_per_fg']:,.6f}".rstrip("0").rstrip(".") if comp["qty_per_fg"] else ""
        gap_badge = '<span class="badge">DATA GAP</span>' if gap else ""
        for mi, (mkey, mlabel) in enumerate(metrics):
            cells = []
            if mi == 0:
                cells.append(
                    f'<td class="comp {cls}" rowspan="{rowspan}">'
                    f'<div class="comp-id">{html.escape(comp["itemid"])} {gap_badge}</div>'
                    f'<div class="comp-desc">{desc}</div>'
                    f'<div class="comp-meta">Current: <b>{cq or "–"}</b> · Qty/FG: {qf or "–"}</div></td>'
                )
            mcls = f"m-{mkey}"
            cells.append(f'<td class="metric {mcls}">{html.escape(mlabel)}</td>')
            for w in weeks:
                v = None if gap else comp["cells"].get(w, {}).get(mkey)
                cells.append(f'<td class="num {mcls}">{_fmt(v)}</td>')
            rcls = "metric-row" + (" emph" if mkey in EMPHASIS else "") + (f" {cls}" if gap else "")
            body_rows.append(f'<tr class="{rcls}">' + "".join(cells) + "</tr>")

    gaps_note = ""
    if model["data_gaps"]:
        gaps_note = (
            '<div class="alert">'
            f'<b>{len(model["data_gaps"])} data gap(s):</b> '
            + ", ".join(html.escape(g) for g in model["data_gaps"])
            + " not found in the planning results. These are shown blank — the value is <b>unknown</b>, not zero."
            "</div>"
        )
    wip = ""
    if model.get("wip_assemblies"):
        names = ", ".join(html.escape(str(w.get("itemid"))) for w in model["wip_assemblies"])
        wip = f'<div class="foot-note">Traversed through {len(model["wip_assemblies"])} WIP sub-assembly(ies) (not shown, made in-house): {names}.</div>'
    nonstock = ""
    if model.get("non_stock_lines"):
        names = ", ".join(
            html.escape(f'{n.get("itemid")} ({n.get("type")})') for n in model["non_stock_lines"]
        )
        nonstock = (
            f'<div class="foot-note">Excluded {len(model["non_stock_lines"])} non-stock / costing BOM line(s) '
            f'(not MRP-planned inventory): {names}.</div>'
        )

    # Legend built from the actual metric set (varies by search variant).
    legend_items = ['<span><b>Rows per item:</b></span>']
    for mkey, mlabel in metrics:
        dot = ""
        if mkey == "planned_demand":
            dot = '<span class="dot" style="background:var(--pd)"></span>'
        elif mkey == "planned_supply":
            dot = '<span class="dot" style="background:var(--ps)"></span>'
        label = f"<b>{html.escape(mlabel)}</b>" if mkey in EMPHASIS else html.escape(mlabel)
        legend_items.append(f"<span>{dot}{label}</span>")
    legend_html = "\n  ".join(legend_items)

    loc = html.escape(model.get("location") or "All locations")
    gen = html.escape(model.get("generated_at") or "")
    fg = html.escape(str(model.get("fg_item") or ""))
    fgd = html.escape(model.get("fg_description") or "")

    doc = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>MRP — {fg}</title>
<style>
:root {{
  --ink:#272838; --fg:#1E293B; --muted:#64748B; --bg:#FFFFFF; --panel:#F5F5F5;
  --well:#F2F2F4; --line:#E0E1E6; --line2:#C1C3CC;
  --pd:#0F9ED5; --ps:#28A745; --end:#272838; --err:#FF0022; --err-bg:#FFE5E9;
  --font:"Assistant",-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;
}}
* {{ box-sizing:border-box; }}
body {{ margin:0; font-family:var(--font); color:var(--fg); background:var(--bg); font-size:13px; }}
.wrap {{ padding:24px 28px 40px; }}
h1 {{ font-size:22px; color:var(--ink); margin:0 0 2px; font-weight:700; }}
.sub {{ color:var(--muted); font-size:12px; margin:0 0 2px; }}
.src {{ color:var(--muted); font-size:11px; margin:0 0 16px; }}
.alert {{ background:var(--err-bg); border:1px solid var(--err); color:#8a0014;
  padding:10px 14px; border-radius:8px; font-size:12px; margin:0 0 16px; }}
.foot-note {{ color:var(--muted); font-size:11px; margin-top:14px; }}
.legend {{ display:flex; gap:16px; font-size:11px; color:var(--muted); margin:0 0 12px; flex-wrap:wrap; }}
.legend b {{ color:var(--fg); }}
.dot {{ display:inline-block; width:9px; height:9px; border-radius:2px; margin-right:4px; vertical-align:middle; }}
.scroller {{ overflow-x:auto; border:1px solid var(--line); border-radius:10px; }}
table {{ border-collapse:separate; border-spacing:0; width:100%; }}
th, td {{ padding:5px 10px; white-space:nowrap; border-bottom:1px solid var(--line); }}
thead th {{ position:sticky; top:0; z-index:3; background:var(--ink); color:#fff; font-weight:600;
  text-align:right; font-size:11px; }}
thead th.corner {{ text-align:left; left:0; z-index:5; }}
thead th.mh {{ text-align:left; left:230px; z-index:5; }}
th.wk .wk-name {{ font-size:11px; }}
th.wk .wk-date {{ font-size:10px; opacity:.72; font-weight:400; }}
td.comp {{ position:sticky; left:0; z-index:2; background:var(--bg); border-right:1px solid var(--line2);
  vertical-align:top; min-width:230px; max-width:230px; white-space:normal; }}
td.metric {{ position:sticky; left:230px; z-index:2; background:var(--bg); border-right:1px solid var(--line2);
  font-size:11px; color:var(--muted); min-width:120px; }}
.comp-id {{ font-weight:700; color:var(--ink); font-size:13px; }}
.comp-desc {{ color:var(--muted); font-size:11px; margin:2px 0; line-height:1.25; }}
.comp-meta {{ font-size:10.5px; color:var(--muted); }}
td.num {{ text-align:right; font-variant-numeric:tabular-nums; }}
.nil {{ color:var(--line2); }}
.m-planned_demand {{ color:var(--pd); }}
.m-planned_supply {{ color:var(--ps); }}
tr.emph td.num, tr.emph td.metric {{ font-weight:700; color:var(--ink); background:#fbfbfd; }}
tr.metric-row:hover td.num {{ background:#eef6fb; }}
tr.gap td, td.comp.gap {{ background:var(--err-bg); }}
.badge {{ background:var(--err); color:#fff; font-size:9px; font-weight:700; padding:1px 5px;
  border-radius:4px; margin-left:4px; vertical-align:middle; }}
</style></head>
<body><div class="wrap">
<h1>Consolidated MRP — purchased components</h1>
<p class="sub">Finished good <b>{fg}</b>{(" — " + fgd) if fgd else ""}</p>
<p class="src">Source: NetSuite saved search {html.escape(model["search_id"])} (Planning Engine Result) · {loc} · generated {gen}</p>
<p class="src">Scope: <b>{len(model["components"])}</b> purchased components · full multi-level BOM explosion · {len(model.get("wip_assemblies") or [])} WIP sub-assembly(ies) traversed (not listed)</p>
{gaps_note}
<div class="legend">
  {legend_html}
</div>
<div class="scroller"><table>
<thead><tr>
  <th class="corner">Component</th>
  <th class="mh">Metric</th>
  {head_cells}
</tr></thead>
<tbody>
{''.join(body_rows)}
</tbody>
</table></div>
{wip}
{nonstock}
<p class="foot-note">Every value traces to saved search {html.escape(model["search_id"])}. Blank = not present in the planning run (unknown, not zero). Purchased components only — the finished good and any WIP sub-assemblies are traversed but not listed.</p>
</div></body></html>"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(doc)


# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--bom", required=True)
    ap.add_argument("--mrp", required=True)
    ap.add_argument("--out-dir", required=True)
    ap.add_argument("--stamp", default="report")
    args = ap.parse_args()

    with open(args.bom, encoding="utf-8") as f:
        bom = json.load(f)
    with open(args.mrp, encoding="utf-8") as f:
        mrp_rows = json.load(f)
    if isinstance(mrp_rows, dict) and "data" in mrp_rows:
        mrp_rows = mrp_rows["data"]

    model = build_model(bom, mrp_rows)
    os.makedirs(args.out_dir, exist_ok=True)
    fg = model["fg_item"] or "FG"
    base = f"MRP_{fg}_{args.stamp}"
    xlsx_path = os.path.join(args.out_dir, base + ".xlsx")
    html_path = os.path.join(args.out_dir, base + ".html")
    write_xlsx(model, xlsx_path)
    write_html(model, html_path)

    print(json.dumps({
        "fg_item": fg,
        "components_rendered": len(model["components"]),
        "data_gaps": model["data_gaps"],
        "weeks": len(model["weeks"]),
        "xlsx": xlsx_path,
        "html": html_path,
    }, indent=2))


if __name__ == "__main__":
    main()
